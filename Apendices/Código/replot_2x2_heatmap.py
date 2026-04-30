"""
Regrafica la matriz 2x2 con el mismo estilo (heatmap) usado en la matriz 8x8.
"""
import os
import numpy as np
import matplotlib.pyplot as plt
import openpyxl

BASE = '/sessions/stoic-affectionate-ritchie/mnt/outputs'

# Cargar valores
wb = openpyxl.load_workbook(os.path.join(BASE, 'matriz_confusion_2x2.xlsx'), data_only=True)
ws = wb['Matriz_2x2']
VP = ws['B4'].value; FN = ws['C4'].value
FP = ws['B5'].value; VN = ws['C5'].value

acc_b = ws.cell(row=10, column=2).value

# Layout estandar:
#                Pred Pos   Pred Neg
# Real Pos        VP         FN
# Real Neg        FP         VN
matrix = np.array([[VP, FN],
                   [FP, VN]])

labels_inner = np.array([['VP', 'FN'],
                         ['FP', 'VN']])
xticks = ['Pred POSITIVO', 'Pred NEGATIVO']
yticks = ['Real POSITIVO', 'Real NEGATIVO']

# Mismo estilo que matriz_confusion.png (8x8): imshow + Blues + texto + colorbar
fig, ax = plt.subplots(figsize=(7, 5.5))
im = ax.imshow(matrix, cmap='Blues')

ax.set_xticks(range(len(xticks)))
ax.set_xticklabels(xticks, fontsize=11, fontweight='bold')
ax.set_yticks(range(len(yticks)))
ax.set_yticklabels(yticks, fontsize=11, fontweight='bold')

# Anotaciones en cada celda: tipo (VP/FN/FP/VN) + valor
for i in range(2):
    for j in range(2):
        v = matrix[i, j]
        color = 'white' if v > matrix.max()/2 else 'black'
        ax.text(j, i-0.12, labels_inner[i, j], ha='center', va='center',
                color=color, fontsize=14, fontweight='bold')
        ax.text(j, i+0.18, str(v), ha='center', va='center',
                color=color, fontsize=22, fontweight='bold')

ax.set_title(f'Matriz de Confusion 2x2 - Ward (k=8)\n'
             f'Micro-promediada one-vs-rest  |  Accuracy binaria = {float(acc_b):.4f}',
             fontsize=12, fontweight='bold')

plt.colorbar(im, ax=ax)
plt.tight_layout()
out = os.path.join(BASE, 'matriz_confusion_2x2.png')
plt.savefig(out, dpi=180, bbox_inches='tight')
plt.close()
print(f'Heatmap 2x2 guardado: {out}')
