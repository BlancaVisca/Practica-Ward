"""
Genera la matriz de confusion 2x2 (VP / VN / FP / FN) a partir del clustering
de 8 grupos con asignacion mediante algoritmo Hungaro.

Metodo: micro-promediado sobre las 8 clases (one-vs-rest sumado).
Para cada par (muestra, clase c):
    VP_c = real == c AND pred == c
    VN_c = real != c AND pred != c
    FP_c = real != c AND pred == c
    FN_c = real == c AND pred != c
La matriz 2x2 final es la suma sobre c=0..7.
"""
import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from scipy.cluster.hierarchy import linkage, fcluster
from scipy.optimize import linear_sum_assignment
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = '/sessions/stoic-affectionate-ritchie/mnt/outputs'

# ---------- Recalcular asignacion (mismo procedimiento que ward_clustering.py) ----------
df = pd.read_csv(os.path.join(BASE, 'dataset_raw', 'DATA (1).csv'))
X = df.drop(columns=['STUDENT ID', 'GRADE']).values.astype(float)
y_true = df['GRADE'].values

Z = linkage(X, method='ward', metric='euclidean')

# Cargar altura del corte de 8 grupos
cortes = pd.read_csv(os.path.join(BASE, 'cortes_resumen.csv'))
h8 = float(cortes.iloc[2]['altura'])
asign = fcluster(Z, t=h8, criterion='distance')
clusters = sorted(set(asign))
classes = sorted(set(y_true))

# Algoritmo Hungaro
cm_raw = np.zeros((len(clusters), len(classes)), dtype=int)
for i, c in enumerate(clusters):
    for j, lab in enumerate(classes):
        cm_raw[i, j] = int(((asign == c) & (y_true == lab)).sum())

n_max = max(len(clusters), len(classes))
cost = np.zeros((n_max, n_max), dtype=int)
cost[:cm_raw.shape[0], :cm_raw.shape[1]] = cm_raw
row_ind, col_ind = linear_sum_assignment(-cost)
cluster_to_class = {clusters[r]: classes[c] for r, c in zip(row_ind, col_ind)
                    if r < len(clusters) and c < len(classes)}

y_pred = np.array([cluster_to_class.get(c, -1) for c in asign])

# ---------- Matriz 2x2 micro-promediada ----------
VP = VN = FP = FN = 0
per_class = []
for c in classes:
    real_pos = (y_true == c)
    pred_pos = (y_pred == c)
    vp = int((real_pos & pred_pos).sum())
    vn = int((~real_pos & ~pred_pos).sum())
    fp = int((~real_pos & pred_pos).sum())
    fn = int((real_pos & ~pred_pos).sum())
    VP += vp; VN += vn; FP += fp; FN += fn
    per_class.append({
        'Clase': f'Clase_{c}',
        'VP (TP)': vp,
        'VN (TN)': vn,
        'FP': fp,
        'FN': fn,
        'Precision': round(vp/(vp+fp), 4) if (vp+fp) else 0.0,
        'Recall': round(vp/(vp+fn), 4) if (vp+fn) else 0.0,
        'F1': round(2*vp/(2*vp+fp+fn), 4) if (2*vp+fp+fn) else 0.0,
    })

print('=== Matriz de confusion 2x2 micro-promediada ===')
print(f'VP (Verdaderos Positivos):  {VP}')
print(f'FN (Falsos Negativos):       {FN}')
print(f'FP (Falsos Positivos):       {FP}')
print(f'VN (Verdaderos Negativos):   {VN}')
print(f'Total observaciones (N x C): {VP+VN+FP+FN}')

# Metricas globales
total = VP + VN + FP + FN
accuracy = (VP + VN) / total
precision = VP / (VP + FP) if (VP+FP) else 0.0
recall = VP / (VP + FN) if (VP+FN) else 0.0
f1 = 2*precision*recall/(precision+recall) if (precision+recall) else 0.0

print(f'\nAccuracy global (binario):   {accuracy:.4f}')
print(f'Precision (micro):            {precision:.4f}')
print(f'Recall (micro):               {recall:.4f}')
print(f'F1 (micro):                   {f1:.4f}')
print(f'\nAccuracy multiclase (=VP/N):  {VP/len(y_true):.4f}  ({VP}/{len(y_true)} aciertos)')

# ---------- Guardar Excel ----------
wb = Workbook()
wb.remove(wb.active)

# Hoja 1: Matriz 2x2
ws = wb.create_sheet('Matriz_2x2')
ws['A1'] = 'Matriz de Confusion 2x2 (Micro-promediada)'
ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
ws['A1'].fill = PatternFill('solid', fgColor='1F3A5F')
ws.merge_cells('A1:D1')
ws['A1'].alignment = Alignment(horizontal='center')

# Etiquetas de cabecera
ws['B3'] = 'Pred POSITIVO'
ws['C3'] = 'Pred NEGATIVO'
ws['A4'] = 'Real POSITIVO'
ws['A5'] = 'Real NEGATIVO'
for c in ['A4','A5','B3','C3']:
    ws[c].font = Font(bold=True, color='FFFFFF')
    ws[c].fill = PatternFill('solid', fgColor='305496')
    ws[c].alignment = Alignment(horizontal='center')

# Valores
ws['B4'] = VP; ws['B4'].fill = PatternFill('solid', fgColor='C6EFCE')
ws['C4'] = FN; ws['C4'].fill = PatternFill('solid', fgColor='FFC7CE')
ws['B5'] = FP; ws['B5'].fill = PatternFill('solid', fgColor='FFC7CE')
ws['C5'] = VN; ws['C5'].fill = PatternFill('solid', fgColor='C6EFCE')

for c in ['B4','C4','B5','C5']:
    ws[c].font = Font(bold=True, size=14)
    ws[c].alignment = Alignment(horizontal='center', vertical='center')

# Etiquetas explicativas
ws['D4'] = 'VP = Verdadero Positivo  |  FN = Falso Negativo'
ws['D5'] = 'FP = Falso Positivo  |  VN = Verdadero Negativo'

# Metricas
ws['A8'] = 'Metricas globales'; ws['A8'].font = Font(bold=True, size=12)
metrics = [
    ('Total observaciones (N x C clases)', total),
    ('Accuracy binaria  = (VP+VN)/Total', round(accuracy,4)),
    ('Precision (micro) = VP/(VP+FP)',    round(precision,4)),
    ('Recall (micro)    = VP/(VP+FN)',    round(recall,4)),
    ('F1 (micro)        = 2*P*R/(P+R)',    round(f1,4)),
    ('Aciertos del clustering (VP)',       VP),
    ('Total estudiantes',                   len(y_true)),
    ('Accuracy multiclase = VP/N',          round(VP/len(y_true),4)),
]
for i, (k, v) in enumerate(metrics, start=9):
    ws.cell(row=i, column=1, value=k).font = Font(bold=True)
    ws.cell(row=i, column=2, value=v)

# Anchos
for col, w in [('A', 38), ('B', 18), ('C', 18), ('D', 50)]:
    ws.column_dimensions[col].width = w

# Hoja 2: detalle por clase
ws2 = wb.create_sheet('Detalle_por_clase')
ws2['A1'] = 'Matriz 2x2 desglosada por clase (one-vs-rest)'
ws2['A1'].font = Font(bold=True, size=12)
headers = ['Clase','VP (TP)','VN (TN)','FP','FN','Precision','Recall','F1']
for i, h in enumerate(headers, start=1):
    cell = ws2.cell(row=3, column=i, value=h)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='305496')
    cell.alignment = Alignment(horizontal='center')

for r, item in enumerate(per_class, start=4):
    for ci, k in enumerate(headers, start=1):
        ws2.cell(row=r, column=ci, value=item[k])

# Fila de totales
total_row = len(per_class) + 4
ws2.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
ws2.cell(row=total_row, column=2, value=VP).font = Font(bold=True)
ws2.cell(row=total_row, column=3, value=VN).font = Font(bold=True)
ws2.cell(row=total_row, column=4, value=FP).font = Font(bold=True)
ws2.cell(row=total_row, column=5, value=FN).font = Font(bold=True)
for c in [2,3,4,5]:
    ws2.cell(row=total_row, column=c).fill = PatternFill('solid', fgColor='FFE699')

for col in 'ABCDEFGH':
    ws2.column_dimensions[col].width = 14

# Hoja 3: Explicacion
ws3 = wb.create_sheet('Explicacion')
explanations = [
    ('Que es la matriz 2x2 en este caso?', None),
    ('', None),
    ('Como el problema es multiclase (8 clases), se construye la matriz 2x2 mediante el', None),
    ('metodo "micro-promediado one-vs-rest": para cada clase c se evalua', None),
    ('cada muestra como problema binario (es de la clase c o no), y se acumulan', None),
    ('VP, VN, FP y FN sobre las 8 clases.', None),
    ('', None),
    ('Definiciones:', None),
    ('  VP (Verdadero Positivo):  real = c  Y  predicho = c', None),
    ('  VN (Verdadero Negativo):  real != c  Y  predicho != c', None),
    ('  FP (Falso Positivo):      real != c  Y  predicho = c', None),
    ('  FN (Falso Negativo):      real = c  Y  predicho != c', None),
    ('', None),
    ('Total de evaluaciones = N estudiantes x C clases = 145 x 8 = 1160', None),
    ('', None),
    ('Nota importante:', None),
    ('En multiclase con asignacion biunivoca cluster->clase (una sola prediccion por', None),
    ('estudiante), se cumple que FP_total = FN_total porque cada error es un FN', None),
    ('para la clase real y un FP para la clase predicha.', None),
]
for i, (txt, _) in enumerate(explanations, start=1):
    ws3.cell(row=i, column=1, value=txt)
ws3.column_dimensions['A'].width = 90

out = os.path.join(BASE, 'matriz_confusion_2x2.xlsx')
wb.save(out)
print(f'\nExcel guardado: {out}')

# ---------- Imagen 2x2 ----------
fig, ax = plt.subplots(figsize=(8, 6))
matrix_2x2 = np.array([[VP, FN], [FP, VN]])
labels_text = np.array([['VP\n(Verdaderos Positivos)', 'FN\n(Falsos Negativos)'],
                         ['FP\n(Falsos Positivos)', 'VN\n(Verdaderos Negativos)']])
colors = np.array([['#A8D8B9', '#F4B7B7'],
                    ['#F4B7B7', '#A8D8B9']])

for i in range(2):
    for j in range(2):
        ax.add_patch(plt.Rectangle((j, 1-i), 1, 1, facecolor=colors[i,j], edgecolor='black', linewidth=1.5))
        ax.text(j+0.5, 1-i+0.65, labels_text[i,j], ha='center', va='center', fontsize=12, fontweight='bold')
        ax.text(j+0.5, 1-i+0.30, str(matrix_2x2[i,j]), ha='center', va='center', fontsize=22, fontweight='bold')

ax.set_xlim(-0.3, 2.3)
ax.set_ylim(-0.3, 2.5)
ax.set_xticks([0.5, 1.5])
ax.set_xticklabels(['Pred POSITIVO', 'Pred NEGATIVO'], fontsize=12, fontweight='bold')
ax.set_yticks([1.5, 0.5])
ax.set_yticklabels(['Real POSITIVO', 'Real NEGATIVO'], fontsize=12, fontweight='bold')
ax.xaxis.tick_top()
ax.set_aspect('equal')
ax.set_title(f'Matriz de Confusion 2x2 (micro-promediada)\n'
             f'Accuracy = {accuracy:.4f}  |  Precision = {precision:.4f}  |  Recall = {recall:.4f}  |  F1 = {f1:.4f}',
             fontsize=12, fontweight='bold', pad=20)
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)
ax.spines['bottom'].set_visible(False)
ax.spines['left'].set_visible(False)
ax.tick_params(length=0)
plt.tight_layout()
plt.savefig(os.path.join(BASE, 'matriz_confusion_2x2.png'), dpi=180, bbox_inches='tight')
plt.close()
print('Imagen guardada: matriz_confusion_2x2.png')
print('\n=== TODO LISTO ===')
