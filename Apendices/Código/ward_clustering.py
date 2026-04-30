"""
Clustering Jerarquico - Metodo de Ward
Dataset: Higher Education Students Performance Evaluation (UCI #856)
Autor: Christopher
Fecha: 2026-04-28

Decisiones:
- Sin estandarizacion (por peticion del usuario)
- Excluye STUDENT ID y GRADE del clustering
- Variables incluidas: 30 preguntas + COURSE ID
- Distancia: euclidiana (requerida por Ward)
"""

import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from scipy.cluster.hierarchy import linkage, dendrogram, fcluster
from scipy.spatial.distance import pdist
from scipy.optimize import linear_sum_assignment
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ---------- Configuracion ----------
BASE = '/sessions/stoic-affectionate-ritchie/mnt/outputs'
DATA_PATH = os.path.join(BASE, 'dataset_raw', 'DATA (1).csv')
os.makedirs(BASE, exist_ok=True)

# ---------- Carga ----------
df = pd.read_csv(DATA_PATH)
print(f'Dataset cargado: {df.shape[0]} estudiantes, {df.shape[1]} columnas')

# Variables para clustering: todo excepto STUDENT ID y GRADE
X = df.drop(columns=['STUDENT ID', 'GRADE']).values.astype(float)
labels_id = df['STUDENT ID'].values
y_true = df['GRADE'].values

print(f'Matriz de clustering: {X.shape}')

# ---------- Ward linkage ----------
Z = linkage(X, method='ward', metric='euclidean')
print(f'Matriz de enlace generada: shape={Z.shape}')
print(f'Rango de distancias Ward: min={Z[:,2].min():.3f}, max={Z[:,2].max():.3f}')

# Guardar matriz de enlace
np.save(os.path.join(BASE, 'linkage_matrix.npy'), Z)

# ---------- Dendrograma ----------
fig, ax = plt.subplots(figsize=(20, 9))
dendrogram(
    Z,
    labels=labels_id,
    leaf_rotation=90,
    leaf_font_size=7,
    color_threshold=0,
    above_threshold_color='#444444',
    ax=ax
)
ax.set_title('Dendrograma - Metodo de Ward\nHigher Education Students Performance Evaluation',
             fontsize=14, fontweight='bold')
ax.set_xlabel('STUDENT ID', fontsize=11)
ax.set_ylabel('Distancia (Ward)', fontsize=11)
ax.grid(axis='y', linestyle='--', alpha=0.4)
plt.tight_layout()
plt.savefig(os.path.join(BASE, 'dendrograma.png'), dpi=180, bbox_inches='tight')
plt.close()
print('Dendrograma guardado: dendrograma.png')

# ---------- Barrido de umbrales para encontrar k=4,6,8 ----------
heights = np.linspace(Z[:,2].min()+0.001, Z[:,2].max(), 500)
ks = []
for h in heights:
    ks.append(fcluster(Z, t=h, criterion='distance').max())
ks = np.array(ks)

scan = pd.DataFrame({'altura': heights, 'k_clusters': ks})
scan.to_csv(os.path.join(BASE, 'barrido_umbrales.csv'), index=False)

def height_for_k(k):
    """Devuelve la altura representativa (punto medio del rango) que produce k clusters."""
    mask = ks == k
    if not mask.any():
        # encontrar k mas cercano disponible
        candidates = sorted(set(ks))
        nearest = min(candidates, key=lambda x: abs(x-k))
        print(f'  ! No existe altura exacta para k={k}, mas cercano k={nearest}')
        mask = ks == nearest
        k = nearest
    rng = heights[mask]
    return float(rng.mean()), int(k)

h4, k4 = height_for_k(4)
h6, k6 = height_for_k(6)
h8, k8 = height_for_k(8)
print(f'Cortes seleccionados:')
print(f'  Corte A: k={k4}, altura={h4:.3f}')
print(f'  Corte B: k={k6}, altura={h6:.3f}')
print(f'  Corte C: k={k8}, altura={h8:.3f}')

# ---------- Dendrograma con lineas de corte ----------
fig, ax = plt.subplots(figsize=(20, 9))
dendrogram(Z, labels=labels_id, leaf_rotation=90, leaf_font_size=7,
           color_threshold=h8, ax=ax)
ax.axhline(h4, color='#1f77b4', linestyle='--', linewidth=2, label=f'Corte A: {k4} grupos (h={h4:.2f})')
ax.axhline(h6, color='#ff7f0e', linestyle='--', linewidth=2, label=f'Corte B: {k6} grupos (h={h6:.2f})')
ax.axhline(h8, color='#d62728', linestyle='--', linewidth=2, label=f'Corte C: {k8} grupos (h={h8:.2f})')
ax.set_title('Dendrograma con Lineas de Corte - Metodo de Ward', fontsize=14, fontweight='bold')
ax.set_xlabel('STUDENT ID', fontsize=11)
ax.set_ylabel('Distancia (Ward)', fontsize=11)
ax.legend(loc='upper right', fontsize=11)
ax.grid(axis='y', linestyle='--', alpha=0.4)
plt.tight_layout()
plt.savefig(os.path.join(BASE, 'dendrograma_con_cortes.png'), dpi=180, bbox_inches='tight')
plt.close()
print('Dendrograma con cortes guardado.')

# ---------- Asignaciones por corte ----------
asign_A = fcluster(Z, t=h4, criterion='distance')
asign_B = fcluster(Z, t=h6, criterion='distance')
asign_C = fcluster(Z, t=h8, criterion='distance')

# ---------- Funcion: generar libro Excel por corte ----------
def generar_libro_corte(asign, k_real, nombre_archivo, label_corte, altura):
    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='305496')

    # hoja resumen
    ws_res = wb.create_sheet('Resumen')
    ws_res['A1'] = f'{label_corte}'
    ws_res['A1'].font = Font(bold=True, size=14)
    ws_res['A2'] = f'Altura de corte: {altura:.4f}'
    ws_res['A3'] = f'Numero de grupos: {k_real}'
    ws_res['A4'] = f'Total estudiantes: {len(asign)}'
    ws_res['A6'] = 'Grupo'
    ws_res['B6'] = 'Tamano'
    ws_res['A6'].font = header_font; ws_res['A6'].fill = header_fill
    ws_res['B6'].font = header_font; ws_res['B6'].fill = header_fill
    for i, g in enumerate(sorted(set(asign)), start=7):
        ws_res.cell(row=i, column=1, value=f'Grupo_{g}')
        ws_res.cell(row=i, column=2, value=int((asign==g).sum()))

    # una hoja por grupo
    for g in sorted(set(asign)):
        ws = wb.create_sheet(f'Grupo_{g}')
        sub = df[asign == g].copy()
        for c, col in enumerate(sub.columns, start=1):
            cell = ws.cell(row=1, column=c, value=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for r, (_, row) in enumerate(sub.iterrows(), start=2):
            for c, val in enumerate(row.values, start=1):
                ws.cell(row=r, column=c, value=val if not isinstance(val, np.integer) else int(val))
        for c in range(1, len(sub.columns)+1):
            ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = 12

    out = os.path.join(BASE, nombre_archivo)
    wb.save(out)
    print(f'  Libro guardado: {nombre_archivo} ({k_real} grupos)')
    return out

print('\nGenerando libros Excel por corte...')
generar_libro_corte(asign_A, k4, f'corte_A_{k4}grupos.xlsx', f'Corte A ({k4} grupos)', h4)
generar_libro_corte(asign_B, k6, f'corte_B_{k6}grupos.xlsx', f'Corte B ({k6} grupos)', h6)
generar_libro_corte(asign_C, k8, f'corte_C_{k8}grupos.xlsx', f'Corte C ({k8} grupos)', h8)

# ---------- Buscar corte mas cercano a 8 grupos ----------
print('\n=== Paso 5: corte para 8 grupos ===')
mask8 = ks == 8
if mask8.any():
    h8_min, h8_max = heights[mask8].min(), heights[mask8].max()
    print(f'Existe rango exacto para k=8: [{h8_min:.4f}, {h8_max:.4f}], altura usada={h8:.4f}')
else:
    print(f'No existe altura exacta para k=8; se usa la mas cercana k={k8}')

# ---------- Matriz de confusion + asignacion humgara ----------
print('\n=== Paso 6: matriz de confusion ===')
asign8 = asign_C  # usar el corte de 8 grupos
clusters = sorted(set(asign8))
classes = sorted(set(y_true))
print(f'Clusters: {len(clusters)}  Clases reales: {len(classes)}')

# matriz cluster x clase
cm_raw = np.zeros((len(clusters), len(classes)), dtype=int)
for i, c in enumerate(clusters):
    for j, lab in enumerate(classes):
        cm_raw[i, j] = int(((asign8 == c) & (y_true == lab)).sum())

cm_raw_df = pd.DataFrame(cm_raw,
                          index=[f'Cluster_{c}' for c in clusters],
                          columns=[f'Clase_{l}' for l in classes])
print('Matriz cluster x clase (sin asignar nombres):')
print(cm_raw_df)

# Algoritmo Hungaro: maximizar la traza (coincidencias)
n_max = max(len(clusters), len(classes))
cost = np.zeros((n_max, n_max), dtype=int)
cost[:cm_raw.shape[0], :cm_raw.shape[1]] = cm_raw
row_ind, col_ind = linear_sum_assignment(-cost)  # negar para maximizar

# Mapear cluster -> clase
cluster_to_class = {}
for r, c in zip(row_ind, col_ind):
    if r < len(clusters) and c < len(classes):
        cluster_to_class[clusters[r]] = classes[c]

print('\nAsignacion (cluster -> clase real):')
for cl in clusters:
    print(f'  Cluster_{cl} -> Clase_{cluster_to_class.get(cl, "N/A")}')

# Predicciones tras la asignacion
y_pred = np.array([cluster_to_class.get(c, -1) for c in asign8])
acc = (y_pred == y_true).mean()
print(f'\nAccuracy global tras asignacion: {acc:.4f} ({(y_pred==y_true).sum()}/{len(y_true)})')

# Matriz de confusion final (clase real x clase predicha)
cm_final = np.zeros((len(classes), len(classes)), dtype=int)
for i, real in enumerate(classes):
    for j, pred in enumerate(classes):
        cm_final[i, j] = int(((y_true == real) & (y_pred == pred)).sum())
cm_final_df = pd.DataFrame(cm_final,
                            index=[f'Real_{l}' for l in classes],
                            columns=[f'Pred_{l}' for l in classes])

# ---------- Guardar matriz de confusion en Excel ----------
wb = Workbook()
wb.remove(wb.active)

ws1 = wb.create_sheet('Matriz_Cluster_vs_Clase')
ws1['A1'] = 'Cluster \\ Clase real'
ws1['A1'].font = Font(bold=True)
for j, cls in enumerate(classes, start=2):
    ws1.cell(row=1, column=j, value=f'Clase_{cls}').font = Font(bold=True)
for i, cl in enumerate(clusters, start=2):
    ws1.cell(row=i, column=1, value=f'Cluster_{cl}').font = Font(bold=True)
    for j, cls in enumerate(classes, start=2):
        ws1.cell(row=i, column=j, value=int(cm_raw[i-2, j-2]))

ws2 = wb.create_sheet('Asignacion_Hungaro')
ws2['A1'] = 'Cluster'; ws2['B1'] = 'Clase asignada'; ws2['C1'] = 'Coincidencias en cluster'; ws2['D1'] = 'Tamano cluster'
for c in 'ABCD':
    ws2[c+'1'].font = Font(bold=True)
for i, cl in enumerate(clusters, start=2):
    cls = cluster_to_class.get(cl)
    ws2.cell(row=i, column=1, value=f'Cluster_{cl}')
    ws2.cell(row=i, column=2, value=f'Clase_{cls}')
    coinc = int(((asign8 == cl) & (y_true == cls)).sum())
    tam = int((asign8 == cl).sum())
    ws2.cell(row=i, column=3, value=coinc)
    ws2.cell(row=i, column=4, value=tam)

ws3 = wb.create_sheet('Matriz_Confusion_Final')
ws3['A1'] = 'Real \\ Pred'
ws3['A1'].font = Font(bold=True)
for j, cls in enumerate(classes, start=2):
    ws3.cell(row=1, column=j, value=f'Pred_{cls}').font = Font(bold=True)
for i, cls in enumerate(classes, start=2):
    ws3.cell(row=i, column=1, value=f'Real_{cls}').font = Font(bold=True)
    for j, cls2 in enumerate(classes, start=2):
        ws3.cell(row=i, column=j, value=int(cm_final[i-2, j-2]))

ws4 = wb.create_sheet('Resumen')
ws4['A1'] = 'Resumen Paso 6 - Matriz de Confusion'; ws4['A1'].font = Font(bold=True, size=14)
ws4['A3'] = 'Numero de clusters'; ws4['B3'] = len(clusters)
ws4['A4'] = 'Numero de clases reales'; ws4['B4'] = len(classes)
ws4['A5'] = 'Aciertos totales'; ws4['B5'] = int((y_pred==y_true).sum())
ws4['A6'] = 'Total estudiantes'; ws4['B6'] = len(y_true)
ws4['A7'] = 'Accuracy global'; ws4['B7'] = round(float(acc), 4)
ws4['A9'] = 'Algoritmo de asignacion'; ws4['B9'] = 'Hungaro (linear_sum_assignment) maximizando coincidencias'

wb.save(os.path.join(BASE, 'matriz_confusion.xlsx'))
print('Matriz de confusion guardada: matriz_confusion.xlsx')

# ---------- Heatmap matriz confusion ----------
fig, ax = plt.subplots(figsize=(9,7))
im = ax.imshow(cm_final, cmap='Blues')
ax.set_xticks(range(len(classes)))
ax.set_xticklabels([f'Pred {c}' for c in classes])
ax.set_yticks(range(len(classes)))
ax.set_yticklabels([f'Real {c}' for c in classes])
for i in range(len(classes)):
    for j in range(len(classes)):
        v = cm_final[i,j]
        ax.text(j, i, v, ha='center', va='center',
                color='white' if v > cm_final.max()/2 else 'black', fontsize=10)
ax.set_title(f'Matriz de Confusion - Ward (k=8)\nAccuracy = {acc:.4f}', fontsize=13, fontweight='bold')
plt.colorbar(im, ax=ax)
plt.tight_layout()
plt.savefig(os.path.join(BASE, 'matriz_confusion.png'), dpi=180, bbox_inches='tight')
plt.close()

# ---------- Guardar resumen de cortes ----------
resumen = pd.DataFrame({
    'corte': ['A','B','C'],
    'k_clusters': [k4, k6, k8],
    'altura': [h4, h6, h8],
})
resumen.to_csv(os.path.join(BASE, 'cortes_resumen.csv'), index=False)

print('\n=== TODO LISTO ===')
print(f'Accuracy final con 8 grupos: {acc:.4f}')
